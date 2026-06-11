"""
Servicio de generación/actualización del fichero Excel.
Usa openpyxl para crear/actualizar el .xlsx con una hoja por mes.
"""

import os
import logging
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from models import Fichaje

logger = logging.getLogger(__name__)

# Ruta del fichero Excel, configurable via variable de entorno
EXCEL_PATH = os.getenv("EXCEL_PATH", "/app/data/fichajes.xlsx")

# ─── Estilos ────────────────────────────────────────────────────────────────

COLOR_CABECERA = "1F4E79"      # Azul oscuro para cabeceras
COLOR_TOTAL = "BDD7EE"         # Azul claro para fila de totales
COLOR_PAUSA = "FFF2CC"         # Amarillo claro para registros tipo pausa
COLOR_FACTURADO = "E2EFDA"     # Verde claro para registros facturados
COLOR_ALTERNO = "F2F2F2"       # Gris muy claro para filas alternas

COLUMNAS = [
    ("Fecha", 12),
    ("Entrada", 10),
    ("Salida", 10),
    ("Horas", 8),
    ("Proyecto", 22),
    ("Cliente", 20),
    ("Tarea", 30),
    ("Pausa", 8),
    ("Facturado", 10),
    ("Notas", 30),
]

def _borde_fino() -> Border:
    """Devuelve un estilo de borde fino para celdas."""
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _estilo_cabecera() -> dict:
    """Estilos para la fila de cabecera."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": PatternFill("solid", fgColor=COLOR_CABECERA),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _borde_fino(),
    }


def _estilo_total() -> dict:
    """Estilos para la fila de totales."""
    return {
        "font": Font(bold=True, size=10),
        "fill": PatternFill("solid", fgColor=COLOR_TOTAL),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": _borde_fino(),
    }


def _aplicar_estilos(cell, estilos: dict):
    """Aplica un diccionario de estilos a una celda."""
    for attr, valor in estilos.items():
        setattr(cell, attr, valor)


def _nombre_hoja(fecha: datetime) -> str:
    """Genera el nombre de la hoja a partir de una fecha: '2025-06'."""
    return fecha.strftime("%Y-%m")


def _formato_horas(horas: float) -> str:
    """Convierte horas decimales a formato 'Xh Ym' (ej: 1.5 → '1h 30m')."""
    if horas is None:
        return ""
    h = int(horas)
    m = round((horas - h) * 60)
    return f"{h}h {m:02d}m"


def exportar_a_excel(fichajes: List[Fichaje]) -> tuple[str, list[str]]:
    """
    Genera o actualiza el fichero Excel con todos los fichajes proporcionados.
    Crea una hoja por cada mes distinto en los datos.
    Devuelve (ruta_del_fichero, lista_de_hojas_generadas).
    """
    # Asegurarse de que el directorio existe
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

    # Cargar el workbook existente o crear uno nuevo
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        logger.info("Actualizando Excel existente en %s", EXCEL_PATH)
    else:
        wb = openpyxl.Workbook()
        # Eliminar la hoja vacía por defecto
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        logger.info("Creando nuevo Excel en %s", EXCEL_PATH)

    # Agrupar fichajes por mes
    por_mes: dict[str, List[Fichaje]] = {}
    for f in fichajes:
        clave = _nombre_hoja(f.entrada)
        por_mes.setdefault(clave, []).append(f)

    hojas_generadas = []

    for mes, registros in sorted(por_mes.items()):
        # Eliminar hoja existente del mes para regenerarla limpia
        if mes in wb.sheetnames:
            del wb[mes]

        ws = wb.create_sheet(title=mes)
        hojas_generadas.append(mes)

        # ─── Cabecera ─────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 30
        for col_idx, (nombre, ancho) in enumerate(COLUMNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nombre)
            _aplicar_estilos(cell, _estilo_cabecera())
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        # Congelar la fila de cabecera
        ws.freeze_panes = "A2"

        # ─── Filas de datos ───────────────────────────────────────────────
        total_horas = 0.0
        for fila_idx, fichaje in enumerate(
            sorted(registros, key=lambda x: x.entrada), start=2
        ):
            es_par = (fila_idx % 2 == 0)
            es_pausa = (fichaje.tipo == "pausa")
            es_facturado = fichaje.facturado

            # Determinar color de fondo de la fila
            if es_pausa:
                color_fondo = COLOR_PAUSA
            elif es_facturado:
                color_fondo = COLOR_FACTURADO
            elif es_par:
                color_fondo = COLOR_ALTERNO
            else:
                color_fondo = "FFFFFF"

            fill = PatternFill("solid", fgColor=color_fondo)

            datos_fila = [
                fichaje.entrada.strftime("%d/%m/%Y"),
                fichaje.entrada.strftime("%H:%M"),
                fichaje.salida.strftime("%H:%M") if fichaje.salida else "",
                _formato_horas(fichaje.horas) if fichaje.horas else "",
                fichaje.proyecto or "",
                fichaje.cliente or "",
                fichaje.tarea or "",
                "Sí" if es_pausa else "No",
                "Sí" if es_facturado else "No",
                fichaje.notas or "",
            ]

            for col_idx, valor in enumerate(datos_fila, start=1):
                cell = ws.cell(row=fila_idx, column=col_idx, value=valor)
                cell.fill = fill
                cell.border = _borde_fino()
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if col_idx in (1, 2, 3, 4, 8, 9) else "left",
                )
                cell.font = Font(size=10)

            if fichaje.horas:
                total_horas += fichaje.horas

        # ─── Fila de totales ──────────────────────────────────────────────
        fila_total = len(registros) + 2
        ws.row_dimensions[fila_total].height = 25
        estilos_total = _estilo_total()

        for col_idx in range(1, len(COLUMNAS) + 1):
            cell = ws.cell(row=fila_total, column=col_idx)
            _aplicar_estilos(cell, estilos_total)

        ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True, size=10)
        _aplicar_estilos(ws.cell(row=fila_total, column=1), estilos_total)

        ws.cell(row=fila_total, column=4, value=_formato_horas(total_horas))
        _aplicar_estilos(ws.cell(row=fila_total, column=4), estilos_total)

        ws.cell(row=fila_total, column=5, value=f"{len(registros)} registros")
        _aplicar_estilos(ws.cell(row=fila_total, column=5), estilos_total)

        logger.info(
            "Hoja '%s' generada con %d registros (%.2f horas totales)",
            mes, len(registros), total_horas
        )

    # Ordenar las hojas por fecha (más reciente al principio)
    wb._sheets.sort(key=lambda ws: ws.title, reverse=True)

    wb.save(EXCEL_PATH)
    logger.info("Excel guardado correctamente en %s", EXCEL_PATH)

    return EXCEL_PATH, hojas_generadas
